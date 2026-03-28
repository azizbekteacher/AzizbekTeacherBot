import asyncio
import io
from datetime import datetime, timedelta

from aiogram import Router, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    Message, CallbackQuery, BotCommand, BotCommandScopeChat,
    ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
    BufferedInputFile,
)
from openpyxl import Workbook

from db import (
    is_admin, add_admin, remove_admin, get_admin_ids,
    get_stats, get_recent_users, get_user_count, get_all_user_ids,
    get_messages_by_category, get_msg, update_msg_text, update_msg_media,
    update_msg_delay, create_custom_message, delete_message, toggle_msg_active,
    get_bookings_by_date, get_week_booking_counts, get_all_users_with_survey,
    get_user_survey_by_telegram_id, get_booking_detail_with_survey,
    get_detailed_stats, get_users_paginated, search_users,
    get_users_without_survey,
)

router = Router()
cmd_router = Router()

USER_COMMANDS = [
    BotCommand(command="start", description="Botni ishga tushirish"),
]

ADMIN_COMMANDS = USER_COMMANDS + [
    BotCommand(command="admins", description="Adminlar ro'yxati"),
    BotCommand(command="addadmin", description="Admin qo'shish"),
    BotCommand(command="removeadmin", description="Admin o'chirish"),
    BotCommand(command="stats", description="Statistika"),
    BotCommand(command="users", description="Foydalanuvchilar"),
    BotCommand(command="search", description="Qidirish"),
    BotCommand(command="broadcast", description="Xabar yuborish"),
    BotCommand(command="messages", description="Bot xabarlari"),
    BotCommand(command="consultations", description="Konsultatsiyalar"),
    BotCommand(command="survey_remind", description="So'rovnoma eslatma yuborish"),
]


class MessageEditor(StatesGroup):
    select_category = State()
    select_message = State()
    view_message = State()
    edit_text = State()
    edit_media = State()
    edit_delay = State()
    new_label = State()
    new_text = State()
    new_media_ask = State()
    new_media = State()
    new_delay = State()


CATEGORY_LABELS = {
    "start": "Start xabarlari",
    "registration": "Ro'yxatdan o'tish",
    "followup": "Follow-up",
    "consultation": "Konsultatsiya",
    "general": "Umumiy",
}


def categories_keyboard() -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(text="Start xabarlari", callback_data="msgcat:start")],
        [InlineKeyboardButton(text="Ro'yxatdan o'tish", callback_data="msgcat:registration")],
        [InlineKeyboardButton(text="Follow-up", callback_data="msgcat:followup")],
        [InlineKeyboardButton(text="Konsultatsiya", callback_data="msgcat:consultation")],
        [InlineKeyboardButton(text="Umumiy", callback_data="msgcat:general")],
    ]
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def messages_list_keyboard(category: str) -> InlineKeyboardMarkup:
    messages = get_messages_by_category(category)
    buttons = []
    for msg in messages:
        status = "" if msg["is_active"] else " [o'chirilgan]"
        buttons.append([InlineKeyboardButton(
            text=f"{msg['label']}{status}",
            callback_data=f"msgview:{msg['key']}",
        )])
    if category == "start":
        buttons.append([InlineKeyboardButton(text="+ Yangi start xabar qo'shish", callback_data="msgcat:new_start")])
    if category == "followup":
        buttons.append([InlineKeyboardButton(text="+ Yangi follow-up qo'shish", callback_data="msgcat:new_followup")])
    buttons.append([InlineKeyboardButton(text="Orqaga", callback_data="msgcat:back")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def message_view_keyboard(msg: dict) -> InlineKeyboardMarkup:
    buttons = [
        [InlineKeyboardButton(text="Matnni o'zgartirish", callback_data=f"msgedit:text:{msg['key']}")],
        [InlineKeyboardButton(text="Media qo'shish/o'zgartirish", callback_data=f"msgedit:media:{msg['key']}")],
    ]
    if msg["category"] in ("followup", "start"):
        if msg["category"] == "followup":
            buttons.append([InlineKeyboardButton(
                text="Vaqtni o'zgartirish", callback_data=f"msgedit:delay:{msg['key']}"
            )])
        active_text = "O'chirish" if msg["is_active"] else "Yoqish"
        buttons.append([InlineKeyboardButton(
            text=active_text, callback_data=f"msgedit:toggle:{msg['key']}"
        )])
    buttons.append([InlineKeyboardButton(text="Orqaga", callback_data=f"msgcat:{msg['category']}")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def format_message_view(msg: dict) -> str:
    cat_label = CATEGORY_LABELS.get(msg["category"], msg["category"])
    lines = [
        f"<b>{msg['label']}</b>",
        f"Kategoriya: {cat_label}",
        f"Turi: {msg['content_type']}",
        "",
    ]
    if msg.get("text"):
        preview = msg["text"][:500]
        if len(msg["text"]) > 500:
            preview += "..."
        lines.append(f"Matn:\n<code>{preview}</code>")
    else:
        lines.append("Matn: yo'q")

    if msg.get("media_file_id"):
        lines.append(f"\nMedia: bor ({msg['content_type']})")
    else:
        lines.append("\nMedia: yo'q")

    if msg["category"] == "followup":
        delay = msg.get("schedule_delay_minutes") or 0
        if delay >= 60:
            delay_text = f"{delay // 60} soat {delay % 60} daqiqa" if delay % 60 else f"{delay // 60} soat"
        else:
            delay_text = f"{delay} daqiqa"
        lines.append(f"Delay: {delay_text}")
        active_label = "ha" if msg["is_active"] else "yo'q"
        lines.append(f"Faol: {active_label}")

    return "\n".join(lines)


# --- /messages command ---


@cmd_router.message(Command("messages"))
async def cmd_messages(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(MessageEditor.select_category)
    await message.answer(
        "<b>Bot xabarlari boshqaruvi</b>\n\nKategoriyani tanlang:",
        reply_markup=categories_keyboard(),
    )


@router.callback_query(MessageEditor.select_category, F.data == "msgcat:back")
async def on_msg_back_to_categories(callback: CallbackQuery, state: FSMContext):
    await state.set_state(MessageEditor.select_category)
    await callback.message.edit_text(
        "<b>Bot xabarlari boshqaruvi</b>\n\nKategoriyani tanlang:",
        reply_markup=categories_keyboard(),
    )
    await callback.answer()


@router.callback_query(MessageEditor.select_category, F.data == "msgcat:new_start")
@router.callback_query(MessageEditor.select_category, F.data == "msgcat:new_followup")
async def on_new_message(callback: CallbackQuery, state: FSMContext):
    cat = "start" if "new_start" in callback.data else "followup"
    await state.update_data(new_category=cat)
    await state.set_state(MessageEditor.new_label)
    cat_label = "start" if cat == "start" else "follow-up"
    await callback.message.edit_text(
        f"<b>Yangi {cat_label} xabar</b>\n\nXabar nomini kiriting (masalan: Eslatma xabar):"
    )
    await callback.answer()


@router.callback_query(MessageEditor.select_category, F.data.startswith("msgcat:"))
async def on_category_selected(callback: CallbackQuery, state: FSMContext):
    category = callback.data.split(":", 1)[1]
    if category not in CATEGORY_LABELS:
        await callback.answer("Noto'g'ri kategoriya", show_alert=True)
        return
    await state.update_data(current_category=category)
    await state.set_state(MessageEditor.select_message)
    cat_label = CATEGORY_LABELS[category]
    await callback.message.edit_text(
        f"<b>{cat_label}</b> xabarlari:\n\nTahrirlash uchun tanlang:",
        reply_markup=messages_list_keyboard(category),
    )
    await callback.answer()


@router.callback_query(MessageEditor.select_message, F.data == "msgcat:new_start")
@router.callback_query(MessageEditor.select_message, F.data == "msgcat:new_followup")
async def on_new_message_from_list(callback: CallbackQuery, state: FSMContext):
    cat = "start" if "new_start" in callback.data else "followup"
    await state.update_data(new_category=cat)
    await state.set_state(MessageEditor.new_label)
    cat_label = "start" if cat == "start" else "follow-up"
    await callback.message.edit_text(
        f"<b>Yangi {cat_label} xabar</b>\n\nXabar nomini kiriting (masalan: Eslatma xabar):"
    )
    await callback.answer()


# Message list navigation — orqaga kategoriyalarga
@router.callback_query(MessageEditor.select_message, F.data == "msgcat:back")
async def on_msg_list_back(callback: CallbackQuery, state: FSMContext):
    await state.set_state(MessageEditor.select_category)
    await callback.message.edit_text(
        "<b>Bot xabarlari boshqaruvi</b>\n\nKategoriyani tanlang:",
        reply_markup=categories_keyboard(),
    )
    await callback.answer()


@router.callback_query(MessageEditor.select_message, F.data.startswith("msgview:"))
async def on_message_selected(callback: CallbackQuery, state: FSMContext):
    key = callback.data.split(":", 1)[1]
    msg = get_msg(key)
    if not msg:
        await callback.answer("Xabar topilmadi", show_alert=True)
        return
    await state.update_data(current_msg_key=key)
    await state.set_state(MessageEditor.view_message)
    await callback.message.edit_text(
        format_message_view(msg),
        reply_markup=message_view_keyboard(msg),
    )
    await callback.answer()


# View message — orqaga kategoriya ro'yxatiga
@router.callback_query(MessageEditor.view_message, F.data.startswith("msgcat:"))
async def on_view_back_to_list(callback: CallbackQuery, state: FSMContext):
    category = callback.data.split(":", 1)[1]
    await state.update_data(current_category=category)
    await state.set_state(MessageEditor.select_message)
    cat_label = CATEGORY_LABELS.get(category, category)
    await callback.message.edit_text(
        f"<b>{cat_label}</b> xabarlari:\n\nTahrirlash uchun tanlang:",
        reply_markup=messages_list_keyboard(category),
    )
    await callback.answer()


# --- Edit text ---
@router.callback_query(MessageEditor.view_message, F.data.startswith("msgedit:text:"))
async def on_edit_text(callback: CallbackQuery, state: FSMContext):
    key = callback.data.split(":", 2)[2]
    await state.update_data(current_msg_key=key)
    await state.set_state(MessageEditor.edit_text)
    await callback.message.edit_text(
        "Yangi matnni yuboring (HTML format).\n\n"
        "Placeholderlar: {full_name}, {phone}, {video_link}, {day_label}, {time_slot}"
    )
    await callback.answer()


@router.message(MessageEditor.edit_text)
async def process_edit_text(message: Message, state: FSMContext):
    data = await state.get_data()
    key = data.get("current_msg_key")
    if not key:
        await state.clear()
        return
    new_text = message.text or message.caption or ""
    update_msg_text(key, new_text)
    msg = get_msg(key)
    await state.set_state(MessageEditor.view_message)
    await message.answer(
        f"Matn yangilandi!\n\n{format_message_view(msg)}",
        reply_markup=message_view_keyboard(msg),
    )


# --- Edit media ---
@router.callback_query(MessageEditor.view_message, F.data.startswith("msgedit:media:"))
async def on_edit_media(callback: CallbackQuery, state: FSMContext):
    key = callback.data.split(":", 2)[2]
    await state.update_data(current_msg_key=key)
    await state.set_state(MessageEditor.edit_media)
    await callback.message.edit_text("Rasm, audio yoki video yuboring:")
    await callback.answer()


@router.message(MessageEditor.edit_media)
async def process_edit_media(message: Message, state: FSMContext):
    data = await state.get_data()
    key = data.get("current_msg_key")
    if not key:
        await state.clear()
        return

    file_id = None
    content_type = "text"

    if message.photo:
        file_id = message.photo[-1].file_id
        content_type = "photo"
    elif message.voice:
        file_id = message.voice.file_id
        content_type = "voice"
    elif message.audio:
        file_id = message.audio.file_id
        content_type = "voice"
    elif message.video:
        file_id = message.video.file_id
        content_type = "video"
    elif message.document:
        file_id = message.document.file_id
        content_type = "document"
    else:
        await message.answer("Iltimos, rasm, audio, video yoki fayl yuboring:")
        return

    if message.caption:
        update_msg_text(key, message.caption)

    update_msg_media(key, file_id, content_type)
    msg = get_msg(key)
    await state.set_state(MessageEditor.view_message)
    await message.answer(
        f"Media saqlandi!\n\n{format_message_view(msg)}",
        reply_markup=message_view_keyboard(msg),
    )


# --- Edit delay ---
@router.callback_query(MessageEditor.view_message, F.data.startswith("msgedit:delay:"))
async def on_edit_delay(callback: CallbackQuery, state: FSMContext):
    key = callback.data.split(":", 2)[2]
    await state.update_data(current_msg_key=key)
    await state.set_state(MessageEditor.edit_delay)
    await callback.message.edit_text(
        "Ro'yxatdan o'tgandan necha daqiqadan keyin yuborilsin?\n\n"
        "Raqam kiriting (masalan: 60 = 1 soat, 1440 = 24 soat):"
    )
    await callback.answer()


@router.message(MessageEditor.edit_delay)
async def process_edit_delay(message: Message, state: FSMContext):
    data = await state.get_data()
    key = data.get("current_msg_key")
    if not key:
        await state.clear()
        return

    text = message.text.strip()
    if not text.isdigit():
        await message.answer("Iltimos, faqat raqam kiriting (daqiqalarda):")
        return

    minutes = int(text)
    update_msg_delay(key, minutes)
    msg = get_msg(key)
    await state.set_state(MessageEditor.view_message)
    await message.answer(
        f"Delay yangilandi!\n\n{format_message_view(msg)}",
        reply_markup=message_view_keyboard(msg),
    )


# --- Toggle active ---
@router.callback_query(MessageEditor.view_message, F.data.startswith("msgedit:toggle:"))
async def on_toggle_active(callback: CallbackQuery, state: FSMContext):
    key = callback.data.split(":", 2)[2]
    toggle_msg_active(key)
    msg = get_msg(key)
    if msg:
        await callback.message.edit_text(
            format_message_view(msg),
            reply_markup=message_view_keyboard(msg),
        )
    await callback.answer("Holat o'zgartirildi!")


# --- New follow-up flow ---
@router.message(MessageEditor.new_label)
async def process_new_label(message: Message, state: FSMContext):
    label = message.text.strip()
    if len(label) < 2:
        await message.answer("Iltimos, kamida 2 belgili nom kiriting:")
        return
    data = await state.get_data()
    cat = data.get("new_category", "followup")
    prefix = "start_custom_" if cat == "start" else "followup_custom_"
    key = prefix + label.lower().replace(" ", "_")[:20]
    await state.update_data(new_key=key, new_label=label)
    await state.set_state(MessageEditor.new_text)
    await message.answer("Xabar matnini yuboring (HTML format):")


@router.message(MessageEditor.new_text)
async def process_new_text(message: Message, state: FSMContext):
    text = message.text or message.caption or ""
    await state.update_data(new_text=text)
    await state.set_state(MessageEditor.new_media_ask)

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="Ha", callback_data="newfu:media_yes"),
            InlineKeyboardButton(text="Yo'q", callback_data="newfu:media_no"),
        ]
    ])
    await message.answer("Media qo'shmoqchimisiz?", reply_markup=kb)


@router.callback_query(MessageEditor.new_media_ask, F.data == "newfu:media_yes")
async def on_new_media_yes(callback: CallbackQuery, state: FSMContext):
    await state.set_state(MessageEditor.new_media)
    await callback.message.edit_text("Rasm, audio yoki video yuboring:")
    await callback.answer()


@router.callback_query(MessageEditor.new_media_ask, F.data == "newfu:media_no")
async def on_new_media_no(callback: CallbackQuery, state: FSMContext):
    await state.update_data(new_file_id=None, new_content_type="text")
    data = await state.get_data()
    if data.get("new_category") == "start":
        await _save_new_message(callback.message, state, delay_minutes=0)
        await callback.answer()
        return
    await state.set_state(MessageEditor.new_delay)
    await callback.message.edit_text(
        "Ro'yxatdan o'tgandan necha daqiqadan keyin yuborilsin?\n\n"
        "Raqam kiriting (masalan: 60 = 1 soat):"
    )
    await callback.answer()


@router.message(MessageEditor.new_media)
async def process_new_media(message: Message, state: FSMContext):
    file_id = None
    content_type = "text"

    if message.photo:
        file_id = message.photo[-1].file_id
        content_type = "photo"
    elif message.voice:
        file_id = message.voice.file_id
        content_type = "voice"
    elif message.audio:
        file_id = message.audio.file_id
        content_type = "voice"
    elif message.video:
        file_id = message.video.file_id
        content_type = "video"
    elif message.document:
        file_id = message.document.file_id
        content_type = "document"
    else:
        await message.answer("Iltimos, rasm, audio, video yoki fayl yuboring:")
        return

    await state.update_data(new_file_id=file_id, new_content_type=content_type)
    data = await state.get_data()
    if data.get("new_category") == "start":
        await _save_new_message(message, state, delay_minutes=0)
        return
    await state.set_state(MessageEditor.new_delay)
    await message.answer(
        "Ro'yxatdan o'tgandan necha daqiqadan keyin yuborilsin?\n\n"
        "Raqam kiriting (masalan: 60 = 1 soat):"
    )


async def _save_new_message(target: Message, state: FSMContext, delay_minutes: int):
    data = await state.get_data()
    await state.clear()

    key = data["new_key"]
    label = data["new_label"]
    category = data.get("new_category", "followup")
    msg_text = data.get("new_text")
    file_id = data.get("new_file_id")
    content_type = data.get("new_content_type", "text")

    ok = create_custom_message(
        key=key,
        label=label,
        category=category,
        text=msg_text,
        content_type=content_type,
        media_file_id=file_id,
        schedule_delay_minutes=delay_minutes if category == "followup" else None,
    )

    from handlers.start import main_menu_kb
    if ok:
        cat_label = CATEGORY_LABELS.get(category, category)
        info = f"Nomi: {label}\nTuri: {content_type}"
        if category == "followup":
            info += f"\nDelay: {delay_minutes} daqiqa"
        await target.answer(
            f"Yangi {cat_label} xabar saqlandi!\n\n{info}",
            reply_markup=main_menu_kb(target.chat.id),
        )
    else:
        await target.answer(
            "Xatolik: Bu nomdagi xabar allaqachon mavjud.",
            reply_markup=main_menu_kb(target.chat.id),
        )


@router.message(MessageEditor.new_delay)
async def process_new_delay(message: Message, state: FSMContext):
    text = message.text.strip()
    if not text.isdigit():
        await message.answer("Iltimos, faqat raqam kiriting (daqiqalarda):")
        return
    await _save_new_message(message, state, delay_minutes=int(text))


class ConsultMsg(StatesGroup):
    waiting_for_content = State()


class Broadcast(StatesGroup):
    waiting_for_content = State()
    confirm = State()


# --- Admin CRUD ---


USERS_PER_PAGE = 10


@cmd_router.message(Command("addadmin"))
async def cmd_add_admin(message: Message):
    if not is_admin(message.from_user.id):
        return

    args = message.text.split(maxsplit=1)
    if len(args) < 2 or not args[1].strip().lstrip("-").isdigit():
        await message.answer(
            "<b>Admin qo'shish</b>\n\n"
            "Foydalanish: <code>/addadmin 123456789</code>\n\n"
            "Telegram ID ni @userinfobot dan olish mumkin."
        )
        return

    new_admin_id = int(args[1].strip())
    if add_admin(new_admin_id, message.from_user.id):
        await message.answer(
            f"<b>Admin qo'shildi!</b>\n\n"
            f"ID: <code>{new_admin_id}</code>\n"
            f"Qo'shgan: <code>{message.from_user.id}</code>"
        )
        try:
            await message.bot.set_my_commands(ADMIN_COMMANDS, scope=BotCommandScopeChat(chat_id=new_admin_id))
        except Exception:
            pass
    else:
        await message.answer(f"Bu foydalanuvchi allaqachon admin: <code>{new_admin_id}</code>")


@cmd_router.message(Command("removeadmin"))
async def cmd_remove_admin(message: Message):
    if not is_admin(message.from_user.id):
        return

    args = message.text.split(maxsplit=1)
    if len(args) < 2 or not args[1].strip().lstrip("-").isdigit():
        await message.answer(
            "<b>Admin o'chirish</b>\n\n"
            "Foydalanish: <code>/removeadmin 123456789</code>"
        )
        return

    target_id = int(args[1].strip())
    if target_id == message.from_user.id:
        await message.answer("O'zingizni adminlikdan chiqara olmaysiz.")
        return

    if remove_admin(target_id):
        await message.answer(
            f"<b>Admin o'chirildi!</b>\n\n"
            f"ID: <code>{target_id}</code>"
        )
        try:
            await message.bot.set_my_commands(USER_COMMANDS, scope=BotCommandScopeChat(chat_id=target_id))
        except Exception:
            pass
    else:
        await message.answer(f"Bu foydalanuvchi admin emas: <code>{target_id}</code>")


@cmd_router.message(Command("admins"))
async def cmd_list_admins(message: Message):
    if not is_admin(message.from_user.id):
        return

    admin_ids = get_admin_ids()
    if not admin_ids:
        await message.answer("Hozircha adminlar yo'q.")
        return

    lines = [f"<b>Adminlar ro'yxati</b> ({len(admin_ids)} ta)\n"]
    for i, aid in enumerate(admin_ids, 1):
        lines.append(f"{i}. <code>{aid}</code>")
    lines.append(
        "\nQo'shish: /addadmin <code>ID</code>"
        "\nO'chirish: /removeadmin <code>ID</code>"
    )
    await message.answer("\n".join(lines))


# --- Stats ---


@cmd_router.message(Command("stats"))
async def cmd_stats(message: Message):
    if not is_admin(message.from_user.id):
        return

    s = get_detailed_stats()
    now = datetime.now().strftime("%d.%m.%Y %H:%M")

    text = (
        f"<b>STATISTIKA</b>\n"
        f"{'─' * 24}\n\n"
        f"<b>Foydalanuvchilar:</b> {s['users_count']} ta\n"
        f"<b>So'rovnomalar:</b> {s['surveys_count']} ta\n"
        f"<b>Adminlar:</b> {s['admins_count']} ta\n\n"
        f"<b>Konsultatsiyalar:</b>\n"
        f"  Jami: {s['bookings_count']} ta\n"
        f"  Bugungi: {s['today_bookings']} ta\n\n"
        f"<b>Scheduled xabarlar:</b>\n"
        f"  Kutilayotgan: {s['pending_msgs']} ta\n"
        f"  Yuborilgan: {s['sent_msgs']} ta\n\n"
        f"{'─' * 24}\n"
        f"Yangilangan: {now}"
    )
    await message.answer(text)


# --- Users (pagination) ---


def users_page_text(users: list[dict], total: int, page: int) -> str:
    total_pages = max(1, (total + USERS_PER_PAGE - 1) // USERS_PER_PAGE)
    lines = [f"<b>Foydalanuvchilar</b> — {total} ta (sahifa {page}/{total_pages})\n"]
    start_num = (page - 1) * USERS_PER_PAGE + 1
    for i, u in enumerate(users, start_num):
        username = u.get("username") or ""
        if username:
            username = f" ({username})"
        lines.append(
            f"{i}. <b>{u['full_name']}</b>{username}\n"
            f"   Tel: {u['phone']}\n"
            f"   ID: <code>{u['telegram_id']}</code>"
        )
    if not users:
        lines.append("(bo'sh)")
    return "\n".join(lines)


def users_page_keyboard(total: int, page: int) -> InlineKeyboardMarkup:
    total_pages = max(1, (total + USERS_PER_PAGE - 1) // USERS_PER_PAGE)
    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton(text="<< Oldingi", callback_data=f"upage:{page - 1}"))
    nav.append(InlineKeyboardButton(text=f"{page}/{total_pages}", callback_data="upage:noop"))
    if page < total_pages:
        nav.append(InlineKeyboardButton(text="Keyingi >>", callback_data=f"upage:{page + 1}"))
    buttons = [nav]
    buttons.append([InlineKeyboardButton(text="Excel yuklab olish", callback_data="cexcel:all")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


@cmd_router.message(Command("users"))
async def cmd_users(message: Message):
    if not is_admin(message.from_user.id):
        return

    total = get_user_count()
    if total == 0:
        await message.answer("Hozircha foydalanuvchilar yo'q.")
        return

    users = get_users_paginated(offset=0, limit=USERS_PER_PAGE)
    text = users_page_text(users, total, page=1)
    kb = users_page_keyboard(total, page=1)
    await message.answer(text, reply_markup=kb)


@router.callback_query(F.data.startswith("upage:"))
async def on_users_page(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return

    page_str = callback.data.split(":", 1)[1]
    if page_str == "noop":
        await callback.answer()
        return

    page = int(page_str)
    if page < 1:
        page = 1

    total = get_user_count()
    offset = (page - 1) * USERS_PER_PAGE
    users = get_users_paginated(offset=offset, limit=USERS_PER_PAGE)
    text = users_page_text(users, total, page)
    kb = users_page_keyboard(total, page)

    try:
        await callback.message.edit_text(text, reply_markup=kb)
    except Exception:
        pass
    await callback.answer()


@cmd_router.message(Command("search"))
async def cmd_search(message: Message):
    if not is_admin(message.from_user.id):
        return

    args = message.text.split(maxsplit=1)
    if len(args) < 2 or len(args[1].strip()) < 2:
        await message.answer(
            "<b>Qidirish</b>\n\n"
            "Foydalanish: <code>/search ism</code>\n"
            "Ism, telefon, ID yoki username bo'yicha qidiradi."
        )
        return

    query = args[1].strip()
    users = search_users(query)
    if not users:
        await message.answer(f"<b>Natija yo'q:</b> \"{query}\"")
        return

    lines = [f"<b>Qidiruv:</b> \"{query}\" — {len(users)} ta natija\n"]
    for i, u in enumerate(users, 1):
        username = u.get("username") or ""
        if username:
            username = f" ({username})"
        lines.append(
            f"{i}. <b>{u['full_name']}</b>{username}\n"
            f"   Tel: {u['phone']}\n"
            f"   ID: <code>{u['telegram_id']}</code>"
        )
    await message.answer("\n".join(lines))


# --- Broadcast ---


@cmd_router.message(Command("cancel"))
async def cmd_cancel(message: Message, state: FSMContext):
    current = await state.get_state()
    if current is None:
        await message.answer("Bekor qilinadigan jarayon yo'q.")
        return
    await state.clear()
    from handlers.start import main_menu_kb
    await message.answer("Bekor qilindi.", reply_markup=main_menu_kb(message.from_user.id))


@cmd_router.message(Command("broadcast"))
async def cmd_broadcast(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return

    total = get_user_count()
    await state.set_state(Broadcast.waiting_for_content)
    await message.answer(
        f"<b>Broadcast</b>\n\n"
        f"Hozir <b>{total}</b> ta foydalanuvchi bor.\n\n"
        "Yubormoqchi bo'lgan xabaringizni yuboring\n"
        "(matn, rasm, video, audio — istalgan turdagi xabar).\n\n"
        "Bekor qilish: /cancel"
    )


@router.message(Broadcast.waiting_for_content)
async def broadcast_content(message: Message, state: FSMContext):
    await state.update_data(
        broadcast_chat_id=message.chat.id,
        broadcast_message_id=message.message_id,
    )
    total = get_user_count()
    await state.set_state(Broadcast.confirm)

    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Test (o'zimga)"), KeyboardButton(text="Ha, yuborish")],
            [KeyboardButton(text="Bekor qilish")]
        ],
        resize_keyboard=True,
        one_time_keyboard=True,
    )
    await message.answer(
        f"Xabar <b>{total}</b> ta foydalanuvchiga yuboriladi.\n\n"
        "Avval o'zingizga test yuboring yoki hammaga yuboring:",
        reply_markup=kb,
    )


@router.message(Broadcast.confirm, F.text == "Bekor qilish")
async def broadcast_cancel(message: Message, state: FSMContext):
    await state.clear()
    from handlers.start import main_menu_kb
    await message.answer("Broadcast bekor qilindi.", reply_markup=main_menu_kb(message.from_user.id))


@router.message(Broadcast.confirm, F.text == "Test (o'zimga)")
async def broadcast_test(message: Message, state: FSMContext):
    data = await state.get_data()
    try:
        await message.bot.copy_message(
            chat_id=message.from_user.id,
            from_chat_id=data["broadcast_chat_id"],
            message_id=data["broadcast_message_id"],
        )
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="Ha, yuborish")],
                [KeyboardButton(text="Bekor qilish")]
            ],
            resize_keyboard=True,
            one_time_keyboard=True,
        )
        await message.answer(
            "Test xabar yuborildi! Yuqoridagi xabarni tekshiring.\n\n"
            "Hammaga yuborasizmi?",
            reply_markup=kb,
        )
    except Exception as e:
        await message.answer(f"Test xatolik: {e}")


@router.message(Broadcast.confirm, F.text == "Ha, yuborish")
async def broadcast_send(message: Message, state: FSMContext):
    data = await state.get_data()
    await state.clear()

    from handlers.start import main_menu_kb

    user_ids = get_all_user_ids()
    sent = 0
    failed = 0
    blocked = 0

    status_msg = await message.answer(
        f"Yuborilmoqda... 0/{len(user_ids)}",
        reply_markup=main_menu_kb(message.from_user.id),
    )

    for i, uid in enumerate(user_ids):
        try:
            await message.bot.copy_message(
                chat_id=uid,
                from_chat_id=data["broadcast_chat_id"],
                message_id=data["broadcast_message_id"],
            )
            sent += 1
        except Exception as e:
            err_msg = str(e).lower()
            if "blocked" in err_msg or "deactivated" in err_msg:
                blocked += 1
            else:
                failed += 1

        if (i + 1) % 25 == 0:
            pct = round((i + 1) / len(user_ids) * 100)
            try:
                await status_msg.edit_text(
                    f"Yuborilmoqda... {i + 1}/{len(user_ids)} ({pct}%)"
                )
            except Exception:
                pass
            await asyncio.sleep(1)

    await status_msg.edit_text(
        f"<b>Broadcast tugadi!</b>\n\n"
        f"Yuborildi: {sent}\n"
        f"Bloklagan: {blocked}\n"
        f"Xatolik: {failed}\n"
        f"{'─' * 24}\n"
        f"Jami: {len(user_ids)}"
    )


# --- So'rovnoma eslatma (survey to'ldirmagan userlarga) ---

SURVEY_REMIND_TEXT = (
    "<b>Assalomu alaykum!</b>\n\n"
    "Siz konsultatsiyaga yozildingiz — bu juda to'g'ri qadam!\n\n"
    "Lekin muhim bir narsa bor:\n\n"
    "Biz siz bilan <b>bog'lanishimiz</b> va konsultatsiyani "
    "<b>samarali o'tkazishimiz</b> uchun sizning telefon raqamingiz "
    "va bir nechta ma'lumotingiz kerak.\n\n"
    "Bu ma'lumotlar orqali:\n"
    "— Sizning darajangizga mos <b>individual reja</b> tayyorlaymiz\n"
    "— Konsultatsiyada vaqt yo'qotmay <b>aniq natijaga</b> chiqamiz\n"
    "— Sizga to'g'ri yo'nalish beramiz\n\n"
    "Atigi <b>2 daqiqa</b> — va biz siz bilan bog'lanamiz!\n\n"
    "Pastdagi tugmani bosing:"
)


@cmd_router.message(Command("survey_remind"))
async def cmd_survey_remind(message: Message):
    if not is_admin(message.from_user.id):
        return

    users = get_users_without_survey()
    if not users:
        await message.answer("Barcha foydalanuvchilar so'rovnomani to'ldirgan.")
        return

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text=f"Ha, {len(users)} ta userga yuborish",
            callback_data="survey_remind:confirm",
        )],
        [InlineKeyboardButton(text="Bekor qilish", callback_data="survey_remind:cancel")],
    ])
    await message.answer(
        f"<b>So'rovnoma eslatma</b>\n\n"
        f"So'rovnoma to'ldirmagan: <b>{len(users)}</b> ta foydalanuvchi\n\n"
        "Ularga so'rovnomani to'ldirish haqida xabar yuborilsinmi?",
        reply_markup=kb,
    )


@router.callback_query(F.data == "survey_remind:cancel")
async def on_survey_remind_cancel(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    await callback.message.edit_text("So'rovnoma eslatma bekor qilindi.")
    await callback.answer()


@router.callback_query(F.data == "survey_remind:confirm")
async def on_survey_remind_confirm(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return

    users = get_users_without_survey()
    if not users:
        await callback.message.edit_text("Barcha foydalanuvchilar allaqachon to'ldirgan.")
        await callback.answer()
        return

    await callback.message.edit_text(f"Yuborilmoqda... 0/{len(users)}")
    await callback.answer()

    survey_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="So'rovnomani to'ldirish", callback_data="survey_fill")],
    ])

    sent = 0
    failed = 0
    blocked = 0

    for i, uid in enumerate(users):
        try:
            await callback.bot.send_message(
                uid, SURVEY_REMIND_TEXT, reply_markup=survey_kb,
            )
            sent += 1
        except Exception as e:
            err_msg = str(e).lower()
            if "blocked" in err_msg or "deactivated" in err_msg:
                blocked += 1
            else:
                failed += 1

        if (i + 1) % 25 == 0:
            pct = round((i + 1) / len(users) * 100)
            try:
                await callback.message.edit_text(
                    f"Yuborilmoqda... {i + 1}/{len(users)} ({pct}%)"
                )
            except Exception:
                pass
            await asyncio.sleep(1)

    await callback.message.edit_text(
        f"<b>So'rovnoma eslatma yuborildi!</b>\n\n"
        f"Yuborildi: {sent}\n"
        f"Bloklagan: {blocked}\n"
        f"Xatolik: {failed}\n"
        f"{'─' * 24}\n"
        f"Jami: {len(users)}"
    )


# --- Konsultatsiyalar ro'yxati ---

WEEKDAY_NAMES = ["Dush", "Sesh", "Chor", "Pay", "Jum", "Shan", "Yak"]
MONTH_NAMES = [
    "", "yanvar", "fevral", "mart", "aprel", "may", "iyun",
    "iyul", "avgust", "sentabr", "oktabr", "noyabr", "dekabr",
]


def format_day_label(date_str: str) -> str:
    d = datetime.strptime(date_str, "%Y-%m-%d").date()
    return f"{WEEKDAY_NAMES[d.weekday()]}, {d.day}-{MONTH_NAMES[d.month]}"


def build_week_keyboard() -> InlineKeyboardMarkup:
    today = datetime.now().date()
    counts = get_week_booking_counts(today.strftime("%Y-%m-%d"), 7)
    buttons = []
    for item in counts:
        d = datetime.strptime(item["date"], "%Y-%m-%d").date()
        label = f"{WEEKDAY_NAMES[d.weekday()]}, {d.day}-{MONTH_NAMES[d.month]} — {item['count']} ta"
        buttons.append([InlineKeyboardButton(
            text=label, callback_data=f"cday:{item['date']}"
        )])
    buttons.append([InlineKeyboardButton(text="Excel yuklab olish (barcha userlar)", callback_data="cexcel:all")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def generate_excel_all_users() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    ws.append([
        "Ism", "Telefon", "Qo'shimcha telefon", "Username", "Telegram ID", "Yosh",
        "Ish/O'qish joyi", "Sinagan usullar", "Kurslar", "Imtihon rejasi",
        "Imtihon maqsadi", "Muhimlik", "Natija ma'nosi", "Byudjet",
        "Video ko'rganmi", "Ro'yxatdan o'tgan sana",
    ])
    users = get_all_users_with_survey()
    for u in users:
        ws.append([
            u["full_name"], u["phone"], u.get("extra_phone") or "",
            u.get("username") or "", u["telegram_id"],
            u.get("age") or "", u.get("workplace") or "",
            u.get("methods_tried") or "", u.get("previous_courses") or "",
            u.get("exam_plan") or "", u.get("exam_goal") or "",
            u.get("importance") or "", u.get("result_meaning") or "",
            u.get("budget") or "", u.get("video_watched") or "",
            u.get("created_at") or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_excel_day_bookings(date_str: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = format_day_label(date_str)
    ws.append([
        "Vaqt", "Ism", "Telefon", "Qo'shimcha telefon", "Username", "Telegram ID",
        "Yosh", "Ish/O'qish joyi", "Sinagan usullar", "Kurslar", "Imtihon rejasi",
        "Imtihon maqsadi", "Muhimlik", "Natija ma'nosi", "Byudjet", "Video ko'rganmi",
    ])
    bookings = get_bookings_by_date(date_str)
    for b in bookings:
        survey = get_user_survey_by_telegram_id(b["telegram_id"])
        s = survey or {}
        ws.append([
            b["time_slot"], b["full_name"], b["phone"], b.get("extra_phone") or "",
            s.get("username") or "", b["telegram_id"],
            s.get("age") or "", s.get("workplace") or "",
            s.get("methods_tried") or "", s.get("previous_courses") or "",
            s.get("exam_plan") or "", s.get("exam_goal") or "",
            s.get("importance") or "", s.get("result_meaning") or "",
            s.get("budget") or "", s.get("video_watched") or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@cmd_router.message(Command("consultations"))
@cmd_router.message(F.text == "Konsultatsiyalar")
async def cmd_consultations(message: Message):
    if not is_admin(message.from_user.id):
        return
    await message.answer(
        "<b>Konsultatsiyalar jadvali</b>\n\n"
        "Kunni tanlang yoki Excel yuklab oling:",
        reply_markup=build_week_keyboard(),
    )


@router.callback_query(F.data == "admin:consultations")
async def on_admin_consultations(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    await callback.message.answer(
        "<b>Konsultatsiyalar jadvali</b>\n\n"
        "Kunni tanlang yoki Excel yuklab oling:",
        reply_markup=build_week_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("cday:"))
async def on_consult_day(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return

    date_str = callback.data.split(":", 1)[1]
    bookings = get_bookings_by_date(date_str)
    day_label = format_day_label(date_str)

    if not bookings:
        await callback.message.edit_text(
            f"<b>{day_label}</b>\n\nBu kunda konsultatsiya yo'q.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Orqaga", callback_data="cweek:back")],
            ]),
        )
        await callback.answer()
        return

    lines = [f"<b>{day_label}</b> — {len(bookings)} ta konsultatsiya\n"]
    for i, b in enumerate(bookings, 1):
        extra = f" | Qo'sh: {b['extra_phone']}" if b.get("extra_phone") else ""
        lines.append(
            f"{i}. <b>{b['time_slot']}</b> — {b['full_name']}\n"
            f"   Tel: {b['phone']}{extra}"
        )

    lines.append("\nBatafsil ko'rish uchun tanlang:")

    buttons = []
    for b in bookings:
        buttons.append([InlineKeyboardButton(
            text=f"{b['time_slot']} — {b['full_name']}",
            callback_data=f"cbuser:{b['booking_id']}",
        )])
    buttons.append([InlineKeyboardButton(text="Excel (shu kun)", callback_data=f"cexcel:{date_str}")])
    buttons.append([InlineKeyboardButton(text="Orqaga", callback_data="cweek:back")])

    await callback.message.edit_text(
        "\n".join(lines),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("cbuser:"))
async def on_consult_user_detail(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return

    booking_id = int(callback.data.split(":", 1)[1])
    detail = get_booking_detail_with_survey(booking_id)

    if not detail:
        await callback.answer("Ma'lumot topilmadi", show_alert=True)
        return

    day_label = format_day_label(detail["date"])

    lines = [
        f"<b>{detail['full_name']}</b>\n",
        f"Konsultatsiya: {day_label}, {detail['time_slot']}",
        f"Tel: {detail['phone']}",
    ]
    if detail.get("extra_phone"):
        lines.append(f"Qo'shimcha tel: {detail['extra_phone']}")

    lines.append(f"Telegram: <code>{detail['telegram_id']}</code>")
    lines.append("")

    fields = [
        ("username", "Username"),
        ("age", "Yosh"),
        ("workplace", "Ish/O'qish"),
        ("methods_tried", "Sinagan usullar"),
        ("previous_courses", "Kurslar"),
        ("exam_plan", "Imtihon rejasi"),
        ("exam_goal", "Imtihon maqsadi"),
        ("importance", "Muhimlik"),
        ("result_meaning", "Natija ma'nosi"),
        ("budget", "Byudjet"),
        ("video_watched", "Video"),
    ]
    for key, label in fields:
        val = detail.get(key)
        if val:
            lines.append(f"<b>{label}:</b> {val}")

    buttons = [
        [InlineKeyboardButton(text="Xabar yuborish", callback_data=f"cbmsg:{detail['telegram_id']}")],
        [InlineKeyboardButton(text="Excel yuklab olish", callback_data=f"cbxl:{detail['telegram_id']}")],
        [InlineKeyboardButton(text="Orqaga", callback_data=f"cday:{detail['date']}")],
    ]

    await callback.message.edit_text(
        "\n".join(lines),
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("cbmsg:"))
async def on_consult_send_msg(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return

    telegram_id = int(callback.data.split(":", 1)[1])
    await state.set_state(ConsultMsg.waiting_for_content)
    await state.update_data(consult_target_id=telegram_id)

    await callback.message.edit_reply_markup(reply_markup=None)
    await callback.message.answer(
        f"Foydalanuvchiga (<code>{telegram_id}</code>) yubormoqchi bo'lgan xabaringizni yuboring.\n\n"
        "Bekor qilish: /cancel",
    )
    await callback.answer()


@router.message(ConsultMsg.waiting_for_content)
async def process_consult_msg(message: Message, state: FSMContext):
    data = await state.get_data()
    target_id = data.get("consult_target_id")
    await state.clear()

    if not target_id:
        return

    try:
        await message.bot.copy_message(
            chat_id=target_id,
            from_chat_id=message.chat.id,
            message_id=message.message_id,
        )
        reply_kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Javob berish", callback_data="user_reply")],
        ])
        await message.bot.send_message(
            target_id,
            "Javob berishingiz mumkin:",
            reply_markup=reply_kb,
        )
        from handlers.start import main_menu_kb
        await message.answer(
            f"Xabar yuborildi! (ID: <code>{target_id}</code>)",
            reply_markup=main_menu_kb(message.from_user.id),
        )
    except Exception as e:
        from handlers.start import main_menu_kb
        await message.answer(
            f"Xabar yuborishda xatolik: {e}",
            reply_markup=main_menu_kb(message.from_user.id),
        )


@router.callback_query(F.data.startswith("cbxl:"))
async def on_consult_user_excel(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return

    telegram_id = int(callback.data.split(":", 1)[1])
    await callback.answer("Excel tayyorlanmoqda...")

    data = generate_excel_single_user(telegram_id)
    if not data:
        await callback.message.answer("Foydalanuvchi topilmadi.")
        return

    doc = BufferedInputFile(data, filename=f"user_{telegram_id}.xlsx")
    await callback.message.answer_document(doc)


def generate_excel_single_user(telegram_id: int) -> bytes | None:
    user = get_user_survey_by_telegram_id(telegram_id)
    if not user:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchi"
    ws.append([
        "Ism", "Telefon", "Qo'shimcha telefon", "Username", "Telegram ID", "Yosh",
        "Ish/O'qish joyi", "Sinagan usullar", "Kurslar", "Imtihon rejasi",
        "Imtihon maqsadi", "Muhimlik", "Natija ma'nosi", "Byudjet",
        "Video ko'rganmi", "Ro'yxatdan o'tgan sana",
    ])
    ws.append([
        user["full_name"], user["phone"], user.get("extra_phone") or "",
        user.get("username") or "", user["telegram_id"],
        user.get("age") or "", user.get("workplace") or "",
        user.get("methods_tried") or "", user.get("previous_courses") or "",
        user.get("exam_plan") or "", user.get("exam_goal") or "",
        user.get("importance") or "", user.get("result_meaning") or "",
        user.get("budget") or "", user.get("video_watched") or "",
        user.get("created_at") or "",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.callback_query(F.data == "cweek:back")
async def on_consult_week_back(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return
    await callback.message.edit_text(
        "<b>Konsultatsiyalar jadvali</b>\n\n"
        "Kunni tanlang yoki Excel yuklab oling:",
        reply_markup=build_week_keyboard(),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("cexcel:"))
async def on_consult_excel(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        await callback.answer()
        return

    target = callback.data.split(":", 1)[1]
    await callback.answer("Excel tayyorlanmoqda...")

    if target == "all":
        data = generate_excel_all_users()
        filename = f"foydalanuvchilar_{datetime.now().strftime('%Y%m%d')}.xlsx"
    else:
        data = generate_excel_day_bookings(target)
        filename = f"konsultatsiya_{target}.xlsx"

    doc = BufferedInputFile(data, filename=filename)
    await callback.message.answer_document(doc)
